import asyncio
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from backend.shared.exceptions import ExportError
from backend.shared.logger import get_logger

logger = get_logger(__name__)

class ExcelExporter:
    """
    Exports stories list to a polished Microsoft Excel workbook.
    """
    async def export(self, stories: List[Dict[str, Any]], output_path: str) -> None:
        """
        Creates worksheets containing structured story matrices.
        """
        logger.info(f"Generating Excel spreadsheet at {output_path}...")
        
        try:
            def build_workbook():
                wb = Workbook()
                # Active worksheet 1: Stories
                ws_stories = wb.active
                ws_stories.title = "User Stories"
                
                # Sheet Styles
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell_font = Font(name="Segoe UI", size=10)
                wrap_align = Alignment(vertical="top", wrap_text=True)
                
                headers = ["Story ID", "Epic", "Feature", "Title", "Story Details", "Acceptance Criteria", "Trace Requirements"]
                ws_stories.append(headers)
                
                for cell in ws_stories[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for story in stories:
                    # Flatten acceptance criteria for text cells
                    ac_texts = []
                    for ac in story.get("acceptance_criteria", []):
                        ac_texts.append(f"Scenario: {ac.get('scenario')}\nGiven {ac.get('given')}\nWhen {ac.get('when')}\nThen {ac.get('then')}")
                    ac_combined = "\n---\n".join(ac_texts)

                    trace_combined = ", ".join(story.get("trace_mappings", []))
                    
                    row = [
                        story.get("id", ""),
                        story.get("epic_id", ""),
                        story.get("feature_id", ""),
                        story.get("title", ""),
                        story.get("user_story_text", ""),
                        ac_combined,
                        trace_combined
                    ]
                    ws_stories.append(row)
                
                # Apply cell styles and alignments
                for r_idx, row in enumerate(ws_stories.iter_rows(min_row=2, max_row=ws_stories.max_row, min_col=1, max_col=7), start=2):
                    for cell in row:
                        cell.font = cell_font
                        cell.alignment = wrap_align
                
                # Set specific column widths for visibility
                ws_stories.column_dimensions["A"].width = 12
                ws_stories.column_dimensions["B"].width = 15
                ws_stories.column_dimensions["C"].width = 15
                ws_stories.column_dimensions["D"].width = 25
                ws_stories.column_dimensions["E"].width = 40
                ws_stories.column_dimensions["F"].width = 40
                ws_stories.column_dimensions["G"].width = 20
                
                wb.save(output_path)
                logger.info("Workbook generated and saved successfully.")

            await asyncio.to_thread(build_workbook)
        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}")
            raise ExportError(f"Excel export failed: {str(e)}")

# INTEGRATION NOTE
# Column auto-wrapping is enabled. Ensure target directory path is writable.
